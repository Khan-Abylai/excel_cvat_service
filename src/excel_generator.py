import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from copy import deepcopy
import os

from src.utils import pixels_to_width_units, get_image_size_with_aspect_ratio, visualize_annotations
import src.config as config

import sys

def generate_excel(images_changed_data, data_folder, cover_all, visualize=True):
    """
    Генерирует Excel-файл с информацией об обработанных изображениях и вставляет обрезанные изображения.
    При необходимости визуализирует аннотации на исходных изображениях.

    :param images_changed_data: Список изображений с успешно обрезанными коробками.
    :param data_folder: Путь к папке с данными.
    :param cover_all: Флаг, указывающий на необходимость покрытия всех данных.
    :param visualize: Флаг, указывающий на необходимость визуализации аннотаций.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    result = []
    max_points = 0
    copied_data = deepcopy(images_changed_data)

    for image_data in copied_data:
        label_groups = {}
        for changed_point in image_data['children']['box']:
            if changed_point['label'] == 'ignore':
                continue
            else:
            # if cover_all is True or changed_point['label'] != 'ignore':
                label = changed_point['label']
                if label not in label_groups:
                    label_groups[label] = []
                label_groups[label].append(changed_point)

        for label, points in label_groups.items():
            image_file_paths = []
            for point in points:
                # Проверяем наличие cropped_name
                cropped_name = point.get('cropped_name')
                if cropped_name:
                    local_file_path = os.path.join(data_folder, 'cropped', cropped_name)
                    image_file_paths.append(local_file_path)
                else:
                    # Если cropped_name отсутствует, добавляем None
                    image_file_paths.append(None)

            # Фильтруем только существующие обрезанные изображения
            valid_image_paths = [p for p in image_file_paths if p is not None]
            if len(valid_image_paths) > max_points:
                max_points = len(valid_image_paths)

            # Формируем строку данных
            column = [
                         label,  # Название
                         None,  # Placeholder для 'фото эталон'
                         None,  # Placeholder для 'task id'
                         image_data['attributes'].get('id', ''),  # ID фото в task
                         image_data['attributes'].get('name', ''),  # Название фото
                         len(valid_image_paths)  # Количество точек на фото
                     ] + valid_image_paths

            result.append(column)

    # Сортируем результаты по названию
    result.sort(key=lambda x: x[0])

    # Формируем заголовки столбцов
    headers = config.headers + [f'фото{x}' for x in range(1, max_points + 1)]
    ws.append(headers)

    # Устанавливаем начальные значения максимальной ширины для изображений
    max_image_widths = [0] * max_points

    # Вставляем данные и изображения в Excel
    for row_idx, data in enumerate(result, start=2):
        max_height_in_row = 0
        for col_idx, cell_value in enumerate(data, start=1):
            if isinstance(cell_value, str) and cell_value.startswith(os.path.join(data_folder, 'cropped')):
                local_file_path = cell_value
                if not os.path.exists(local_file_path):
                    print(f"Предупреждение: файл {local_file_path} не найден. Пропуск вставки изображения.")
                    continue
                try:
                    img_width, img_height = get_image_size_with_aspect_ratio(local_file_path, config.img_max_width)
                    img = ExcelImage(local_file_path)
                    img.width, img.height = img_width, img_height
                    max_height_in_row = max(max_height_in_row, img.height)
                    ws.add_image(img, f"{get_column_letter(col_idx)}{row_idx}")

                    # Обновляем максимальную ширину столбца
                    image_col_index = col_idx - config.headers_count
                    if 0 <= image_col_index < len(max_image_widths):
                        max_image_widths[image_col_index] = max(max_image_widths[image_col_index], img_width)
                except Exception as e:
                    print(f"Ошибка при вставке изображения {local_file_path} в Excel: {e}")
            else:
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Устанавливаем высоту строки на основе максимальной высоты вставленных изображений
        if max_height_in_row > 0:
            ws.row_dimensions[row_idx].height = max_height_in_row * config.row_height_coef

    # Визуализируем аннотации, если необходимо
    if visualize:
        visualization_folder = os.path.join(data_folder, 'rec')
        os.makedirs(visualization_folder, exist_ok=True)
        for image_data in copied_data:
            image_name = image_data['attributes'].get('name', '')
            image_path = os.path.join(data_folder, 'images', image_name)
            annotations = image_data.get('children', {})
            output_image_filename = f"annotated_{image_name}"
            output_image_path = os.path.join(visualization_folder, output_image_filename)
            visualize_annotations(image_path, annotations, visualization_folder, output_image_filename)

    # Устанавливаем ширину столбцов для изображений
    for i, max_width in enumerate(max_image_widths, start=config.headers_count):
        if max_width > 0:
            ws.column_dimensions[get_column_letter(i)].width = pixels_to_width_units(max_width)

    # Устанавливаем фиксированные ширины для столбцов 'A' и 'E'
    ws.column_dimensions['A'].width = config.column_a_width
    ws.column_dimensions['E'].width = config.column_e_width

    # Сохраняем Excel-файл
    output_path = os.path.join(data_folder, 'output.xlsx')
    wb.save(output_path)

    # Удаляем точки из 'images_changed_data' после визуализации и сохранения Excel
    for image_data in images_changed_data:
        children = image_data.get('children', {})
        boxes = children.get('box', [])
        # Удаляем коробки с меткой 'ignore'
        image_data['children']['box'] = [x for x in boxes if x.get('label', '').lower() != 'ignore']
        # Удаляем точки
        image_data['children'].pop('points', None)

